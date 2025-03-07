from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify, session, flash
from flask_sqlalchemy import SQLAlchemy
from flask_bootstrap import Bootstrap
from flask_bcrypt import Bcrypt
from datetime import datetime
from sqlalchemy import extract, func

import calendar
import pandas as pd
from functools import wraps

from forms import Login

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Border, Side
from io import BytesIO

app = Flask(__name__)

app.config['SECRET_KEY'] = '$#PGasdm23'
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root@localhost/aplikasipencatatan'
app.config['SQLALCHEMY_TRACK_MODIFICATION'] = True

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
bootstrap = Bootstrap(app)

class Admin(db.Model):
    __tablename__ = 'admin'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(10))
    password = db.Column(db.Text)

    def __init__(self, username, password):
        self.username = username
        if password != '':
            self.password = bcrypt.generate_password_hash(password). decode('UTF-8')
        

class Order(db.Model):
    __tablename__ = 'order'
    no_pesanan = db.Column(db.Integer, primary_key=True)
    nama_pemesan = db.Column(db.String(20))
    quantity = db.Column(db.Float)
    jelly = db.Column(db.String(5))
    alamat = db.Column(db.Text)
    total_harga = db.Column(db.Float)
    status = db.Column(db.String(15))
    tanggal = db.Column(db.Date)

    def __init__(self, nama_pemesan, quantity, jelly, alamat, total_harga, status, tanggal):
        self.nama_pemesan = nama_pemesan
        self.quantity = quantity
        self.jelly = jelly
        self.alamat = alamat
        self.total_harga = total_harga
        self.status = status
        self.tanggal = tanggal

db.create_all()

def login_required(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if 'login' in session:
            return f(*args, **kwargs)
        else:
            return redirect(url_for('login'))
    return wrap

# Filter template (fungsi untuk mata uang)
@app.template_filter('currency_format')
def currency_format(value):
    value = float(value)
    return "Rp. {:,.0f}".format(value)

#Filter untuk satuan
@app.template_filter('quantity_format')
def quantity_format(value):
    try:
        value = float(value)  # Konversi nilai menjadi float
        return "{:,.0f} pcs".format(value)  # Format tanpa angka desimal, tambahkan "PCS"
    except (ValueError, TypeError):
        return "0 PCS"  # Tanggulangi jika value tidak valid
    
@app.template_filter('month_name')
def month_name_filter(month):
    return calendar.month_name[month]

@app.route('/')
def index():
    if session.get('login') == True:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    # get bulan dan tahun saat ini
    month = datetime.now().month
    year = datetime.now().year

    # perhitungan produk terjual bulan saat ini
    monthly_product = db.session.query(func.sum(Order.quantity)).filter(
        db.extract('month', Order.tanggal) == month,
        db.extract('year', Order.tanggal) == year
    ).scalar() or 0

    # Perhitunhan produk terjual secara keseluruhan
    total_product = db.session.query(func.sum(Order.quantity)).scalar() or 0

    # Perhitungan penghasilan bulan saat ini
    monthly_income = db.session.query(func.sum(Order.total_harga)).filter(
        db.extract('month', Order.tanggal) == month,
        db.extract('year', Order.tanggal) == year
    ).scalar() or 0

    # Perhitungan penghasilan secara keseluruhan
    total_income = db.session.query(func.sum(Order.total_harga)).scalar() or 0

    return render_template('main.html', monthly_product=monthly_product, total_product=total_product, 
                           monthly_income=monthly_income, total_income=total_income)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('login') == True:
        return redirect(url_for('dashboard'))
    forms = Login()
    if forms.validate_on_submit():
        admin = Admin.query.filter_by(username=forms.username.data). first()
        if admin:
            if bcrypt.check_password_hash(admin.password, forms.password.data):
                session['login'] = True
                session['id'] = admin.id
                session['username'] = admin.username
                return redirect(url_for('dashboard'))
        error = "Data yang anda masukkan salah"
        return render_template('login.html', error=error, forms=forms)
    return render_template('login.html', forms=forms)


@app.route('/api/monthly-data', methods=['GET'])
def monthly_data():
    # Ambil tahun dari parameter request, atau gunakan tahun saat ini
    tahun = request.args.get('tahun', datetime.now().year, type=int)

    # Query data total produk berdasarkan bulan
    monthly_data = db.session.query(
        db.extract('month', Order.tanggal).label('bulan'),
        func.sum(Order.quantity).label('total_produk')
    ).filter(
        db.extract('year', Order.tanggal) == tahun
    ).group_by(
        db.extract('month', Order.tanggal)
    ).order_by(
        db.extract('month', Order.tanggal)
    ).all()

    # Format data sebagai JSON
    data = [{'bulan': int(row.bulan), 'total_produk': row.total_produk or 0} for row in monthly_data]

    return jsonify(data)

@app.route('/order')
@login_required
def order():
    query = request.args.get('search')
    if query:
        orders = Order.query.filter(
            Order.status == "Dalam Proses",
            (
                Order.nama_pemesan.ilike(f"%{query}%") |
                Order.alamat.ilike(f"%{query}%")
            )
        ).all()
    else:
        orders = Order.query.filter_by(status="Dalam Proses").all()
    return render_template('order.html', orders=orders)

@app.route('/tambahorder', methods=['GET', 'POST'])
@login_required
def tambah_order():
    if request.method == "POST":
        nama_pemesan = request.form['nama_pemesan']
        quantity = float(request.form['quantity'])
        jelly = request.form['jelly']
        alamat = request.form['alamat']

        total_harga = quantity * 8000
        status = "Dalam Proses"
        tanggal = datetime.now().date()
        db.session.add(Order(nama_pemesan, quantity, jelly, alamat, total_harga, status, tanggal))
        db.session.commit()
        return redirect(url_for('order'))

@app.route('/selesaiorder/<no_pesanan>', methods=['GET', 'POST'])
def selesai_order(no_pesanan):
    order = Order.query.filter_by(no_pesanan=no_pesanan). first()
    if request.method == "POST":
        order.status = "Selesai"
        db.session.commit()
    return redirect(request.referrer)


@app.route('/hapusorder/<no_pesanan>', methods=['GET', 'POST'])
def hapus_order(no_pesanan):
    order = Order.query.filter_by(no_pesanan=no_pesanan). first()
    db.session.delete(order)
    db.session.commit()
    return redirect(request.referrer)

@app.route('/report')
@login_required
def report():
    orders = Order.query.all()

    month_year = db.session.query(
        extract('month', Order.tanggal).label('bulan'),
        extract('year', Order.tanggal).label('tahun')
    ).distinct().all()

    return render_template('report.html', orders=orders, month_year=month_year)

@app.route('/monthreport', methods=['POST'])
def month_report():
    bulan = request.form['bulan']
    tahun = request.form['tahun']

    orders = Order.query.filter(
        db.extract('month', Order.tanggal) == bulan,
        db.extract('year', Order.tanggal) == tahun
    ).all()

    if not orders:
        flash(f"Tidak ada data pesanan untuk bulan {bulan} tahun {tahun}.", "warning")
        return redirect(request.referrer or url_for('report'))

    # data order menjadi dataframe
    data = [{
        "No Pesanan": order.no_pesanan,
        "Nama Pemesan": order.nama_pemesan,
        "Quantity": order.quantity,
        "Jelly": order.jelly,
        "Alamat": order.alamat,
        "Total Harga": order.total_harga,
        "Status": order.status,
        "Tanggal": order.tanggal
    } for order in orders]

    df = pd.DataFrame(data)

    # Hitung total uang dan total produk
    total_uang = df["Total Harga"].sum()
    total_produk = df["Quantity"].sum()

    # Buat workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"

    # Header laporan
    ws.append([f"Laporan Penjualan Bulan {bulan} Tahun {tahun}"])
    ws.append([])
    ws.append(["Total Uang", ":", f"Rp {total_uang:,.0f}"])
    ws.append(["Total Produk Terjual", ":", f"{total_produk:,.0f} pcs"])
    ws.append([])

    # Header tabel data
    headers = list(df.columns)
    ws.append(headers)

    # Masukkan data dari DataFrame ke dalam sheet Excel
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    
    # Styling
    # Header laporan
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:E1")  # Gabungkan sel untuk header laporan

    # Header tabel
    header_row = ws[6]  # Baris ke-5 adalah header tabel
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Simpan laporan ke buffer (untuk diunduh)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Laporan_Penjualan_{bulan}_{tahun}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/yearreport', methods=['POST'])
def year_report():
    # Ambil data tahun dari form
    tahun = request.form['tahun']

    # Query data berdasarkan tahun saja
    orders = Order.query.filter(
        db.extract('year', Order.tanggal) == tahun
    ).all()

    # Periksa apakah ada data untuk tahun yang diminta
    if not orders:
        flash(f"Tidak ada data pesanan untuk tahun {tahun}.", "warning")
        return redirect(request.referrer or url_for('report'))

    # Data order menjadi dataframe
    data = [{
        "No Pesanan": order.no_pesanan,
        "Nama Pemesan": order.nama_pemesan,
        "Quantity": order.quantity,
        "Jelly": order.jelly,
        "Alamat": order.alamat,
        "Total Harga": order.total_harga,
        "Status": order.status,
        "Tanggal": order.tanggal
    } for order in orders]

    df = pd.DataFrame(data)

    # Hitung total uang dan total produk
    total_uang = df["Total Harga"].sum()
    total_produk = df["Quantity"].sum()

    # Buat workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"

    # Header laporan
    ws.append([f"Laporan Penjualan Tahun {tahun}"])
    ws.append([])
    ws.append(["Total Uang", ":", f"Rp {total_uang:,.0f}"])
    ws.append(["Total Produk Terjual", ":", f"{total_produk:,.0f} pcs"])
    ws.append([])

    # Header tabel data
    headers = list(df.columns)
    ws.append(headers)

    # Masukkan data dari DataFrame ke dalam sheet Excel
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    # Styling
    # Header laporan
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:E1")  # Gabungkan sel untuk header laporan

    # Header tabel
    header_row = ws[6]  # Baris ke-6 adalah header tabel
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Simpan laporan ke buffer (untuk diunduh)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Laporan_Penjualan_{tahun}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/logout')
@login_required
def logout():
    session.clear()
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True)